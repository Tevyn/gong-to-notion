"""One-off: populate the `Account ID` property on Agencies from an SFDC xlsx.

Matches rows in `Tevyn's Accounts Report-*.xlsx` to Notion Agencies by
normalized Account Name, and writes the 15-char Salesforce Account ID into
the Agency's `Account ID` property when it's blank. Never overwrites a
non-empty value — conflicts are logged, not clobbered.

Run:
    uv run --with openpyxl python fill_agency_account_ids.py PATH.xlsx
    uv run --with openpyxl python fill_agency_account_ids.py PATH.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv

from src.gong_to_notion.agency_and_staff_fill import (
    AGENCIES_DATA_SOURCE_ID,
    normalize_sf_account_id,
)
from src.gong_to_notion.notion_client import NotionClient

_ROOT = Path(__file__).resolve().parent

PROPERTY_NAME = "Account ID"
HEADER_SENTINEL = "Account Name"  # the row that starts the data table in the xlsx


# ---------------------------------------------------------------------------
# Name normalization — matches xlsx Account Name to Notion Agency title
# ---------------------------------------------------------------------------

_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")
_TRAILING_PAREN_RE = re.compile(r"\s*\(([^()]*)\)\s*$")


def normalize_name(raw: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    s = (raw or "").strip().lower()
    s = _PUNCT_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip()
    return s


def name_candidates(raw: str) -> set[str]:
    """Return normalized candidate keys for a given raw name.

    For `Foo Bar (Baz)` we generate: the full form, the part before the
    parenthetical, and the parenthetical content itself. This lets xlsx
    rows match Notion agencies whether the short alias lives in the
    `Name` or in a trailing parenthetical on either side.
    """
    keys: set[str] = set()
    s = (raw or "").strip()
    if not s:
        return keys
    full = normalize_name(s)
    if full:
        keys.add(full)
    m = _TRAILING_PAREN_RE.search(s)
    if m:
        outer = _TRAILING_PAREN_RE.sub("", s).strip()
        inner = m.group(1).strip()
        if outer:
            k = normalize_name(outer)
            if k:
                keys.add(k)
        if inner:
            k = normalize_name(inner)
            if k:
                keys.add(k)
    return keys


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------


def parse_accounts_xlsx(path: Path) -> list[tuple[str, str]]:
    """Return [(original_name, sf_id_15), ...] deduped by xlsx Account Name.

    The report has leading title/filter rows; we find the data table by
    locating the header row whose first non-empty cell is 'Account Name'.
    """
    import openpyxl  # type: ignore[import-not-found]

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx: int | None = None
    header: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if HEADER_SENTINEL in cells:
            header_row_idx = i
            header = cells
            break
    if header_row_idx is None:
        raise SystemExit(
            f"Could not find header row containing {HEADER_SENTINEL!r} in {path}"
        )

    try:
        name_col = header.index("Account Name")
        id_col = header.index("Account ID")
    except ValueError as e:
        raise SystemExit(f"Missing expected column in header: {e}")

    out: list[tuple[str, str]] = []
    seen_names: set[str] = set()
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = (row[name_col] if name_col < len(row) else None) or ""
        raw_id = (row[id_col] if id_col < len(row) else None) or ""
        name = str(name).strip()
        sf_id = normalize_sf_account_id(str(raw_id).strip())
        if not name or not sf_id:
            continue
        if name in seen_names:
            print(
                f"[xlsx] WARN: duplicate Account Name {name!r} — keeping first",
                file=sys.stderr,
            )
            continue
        seen_names.add(name)
        out.append((name, sf_id))
    return out


# ---------------------------------------------------------------------------
# Notion read helpers (local — keeps this script self-contained)
# ---------------------------------------------------------------------------


def _read_title(prop: dict | None) -> str:
    if not prop:
        return ""
    parts = prop.get("title") or []
    return "".join((p.get("plain_text") or "") for p in parts).strip()


def _read_rich_text(prop: dict | None) -> str:
    if not prop:
        return ""
    parts = prop.get("rich_text") or []
    return "".join((p.get("plain_text") or "") for p in parts).strip()


def _rich_text_payload(value: str) -> dict:
    return {"rich_text": [{"type": "text", "text": {"content": value}}]}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    load_dotenv(_ROOT / ".env")
    token = os.getenv("NOTION_TOKEN")
    if not token:
        raise SystemExit("NOTION_TOKEN must be set in .env")

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to the SFDC Accounts Report xlsx.")
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Report the planned writes without modifying Notion.",
    )
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"File not found: {args.xlsx}")

    print(f"[xlsx] reading {args.xlsx} ...", file=sys.stderr)
    xlsx_rows = parse_accounts_xlsx(args.xlsx)
    print(f"  {len(xlsx_rows)} (name → SF ID) pairs", file=sys.stderr)

    with NotionClient(token) as notion:
        print("[notion] loading Agencies ...", file=sys.stderr)
        agencies = notion.query_data_source(AGENCIES_DATA_SOURCE_ID)
        print(f"  {len(agencies)} agency pages", file=sys.stderr)

        # Auto-detect the Account ID property type from the first page that has it.
        detected_type: str | None = None
        for a in agencies:
            prop = (a.get("properties") or {}).get(PROPERTY_NAME)
            if prop:
                detected_type = prop.get("type")
                break
        if detected_type is None:
            raise SystemExit(
                f"No Agency page exposes a {PROPERTY_NAME!r} property — is the "
                "property actually on the Agencies database and shared with the "
                "integration?"
            )
        if detected_type != "rich_text":
            raise SystemExit(
                f"Unexpected type for {PROPERTY_NAME!r}: {detected_type!r}. "
                "This script only knows how to write to a rich_text (plain text) "
                "property. Change the property type or extend the script."
            )

        # Index Notion pages by every candidate key we can derive from Name
        # and Full name (full form, leading segment, and trailing-parenthetical
        # alias). Each page may register under multiple keys; one key may map
        # to multiple pages (ambiguous).
        key_to_pages: dict[str, set[str]] = {}
        page_info: dict[str, tuple[str, str, str]] = {}
        # page_info[page_id] = (display_name, current_account_id, full_name)
        for a in agencies:
            pid = a.get("id")
            if not pid:
                continue
            props = a.get("properties") or {}
            name = _read_title(props.get("Name"))
            full = _read_rich_text(props.get("Full name"))
            current = _read_rich_text(props.get(PROPERTY_NAME))
            page_info[pid] = (name, current, full)
            keys = name_candidates(name) | name_candidates(full)
            for k in keys:
                key_to_pages.setdefault(k, set()).add(pid)

        filled: list[tuple[str, str]] = []          # (name, sf_id)
        already_set: list[tuple[str, str]] = []     # (name, sf_id)
        conflicts: list[tuple[str, str, str]] = []  # (name, xlsx_id, notion_id)
        ambiguous: list[tuple[str, list[str]]] = [] # (xlsx_name, [notion_names])
        unmatched_xlsx: list[tuple[str, str]] = []
        matched_pages: set[str] = set()

        for xlsx_name, sf_id in xlsx_rows:
            candidates = name_candidates(xlsx_name)
            hits: set[str] = set()
            for k in candidates:
                hits |= key_to_pages.get(k, set())
            if not hits:
                unmatched_xlsx.append((xlsx_name, sf_id))
                continue
            if len(hits) > 1:
                ambiguous.append(
                    (xlsx_name, sorted(page_info[pid][0] for pid in hits))
                )
                continue
            (pid,) = hits
            matched_pages.add(pid)
            notion_name, current, _ = page_info[pid]
            if current:
                if normalize_sf_account_id(current) == sf_id:
                    already_set.append((notion_name, sf_id))
                else:
                    conflicts.append((notion_name, sf_id, current))
                continue
            if args.dry_run:
                filled.append((notion_name, sf_id))
                continue
            notion.update_page(
                pid, {PROPERTY_NAME: _rich_text_payload(sf_id)}
            )
            filled.append((notion_name, sf_id))
            print(f"  wrote {sf_id} → {notion_name!r}", file=sys.stderr)

        unmatched_agencies = [
            (info[0], info[1])
            for pid, info in page_info.items()
            if pid not in matched_pages
        ]

    # -------- Report --------
    print("")
    print("=" * 72)
    print(f"{'DRY-RUN — ' if args.dry_run else ''}Agency Account ID backfill")
    print("=" * 72)
    print(f"  {'Would fill' if args.dry_run else 'Filled'}:     {len(filled)}")
    print(f"  Already set:     {len(already_set)}")
    print(f"  Conflicts:       {len(conflicts)}")
    print(f"  Ambiguous:       {len(ambiguous)}  (xlsx name matches multiple Agencies)")
    print(f"  Unmatched xlsx:  {len(unmatched_xlsx)}  (no Agency with that name)")
    print(
        f"  Unmatched pages: {len(unmatched_agencies)}  "
        f"(Agencies not referenced by the xlsx)"
    )

    if conflicts:
        print("\nConflicts (existing Account ID differs from xlsx — NOT overwritten):")
        for name, xlsx_id, notion_id in conflicts:
            print(f"  - {name!r}: xlsx={xlsx_id}  notion={notion_id}")

    if ambiguous:
        print("\nAmbiguous matches (resolve by hand):")
        for xlsx_name, notion_names in ambiguous:
            print(f"  - {xlsx_name!r} ↔ {notion_names}")

    if unmatched_xlsx:
        print("\nxlsx rows with no matching Agency page:")
        for name, sf_id in sorted(unmatched_xlsx):
            print(f"  - {name}  ({sf_id})")

    if unmatched_agencies:
        print("\nAgency pages with no xlsx row (no Account ID assigned):")
        for name, current in sorted(unmatched_agencies):
            suffix = f"  [already has {current}]" if current else ""
            print(f"  - {name}{suffix}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
