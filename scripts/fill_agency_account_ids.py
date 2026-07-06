"""Upsert Notion Agencies from an SFDC Accounts Report xlsx.

Matches xlsx rows to existing Notion Agencies by SF Account ID (preferred)
or normalized Account Name (alias-aware), and:
  - On matched pages, fills any of the supported properties that are blank
    in Notion. With --update, also overwrites values that differ from the
    xlsx (SF is the source of truth); xlsx blanks never clear a Notion value.
    The page title is never written on matched pages — mismatches between
    the Notion title and the SF Account Name are reported only.
  - On unmatched xlsx rows, optionally CREATES a new Agency page with
    Name, Account ID, Account Stage, Classification, and any other
    supported columns present in the row. --create-stages limits creation
    to rows whose Account Stage is in the given comma-separated list.
  - Derives the Salesforce Link URL from the Account ID.
  - Links Parent Agency from the "Parent Account ID" column in a second
    pass (so parents created in the same run resolve).

Default mode is dry-run; pass --apply to actually write. Pass --no-create
to skip creating new pages (updates only). Without --update, conflicts
(existing value differs from xlsx) are logged, never clobbered.

Run (from the repo root; needs the `scripts` extra for openpyxl):
    uv run --extra scripts python scripts/fill_agency_account_ids.py PATH.xlsx
    uv run --extra scripts python scripts/fill_agency_account_ids.py PATH.xlsx --update
    uv run --extra scripts python scripts/fill_agency_account_ids.py PATH.xlsx --update --apply --create-stages "Won,Engaged"
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

# This script lives in scripts/; add the repo root to sys.path so the
# `src...` imports resolve when run as `uv run python scripts/<name>.py`.
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.gong_to_notion.agency_and_staff_fill import (
    AGENCIES_DATA_SOURCE_ID,
    normalize_sf_account_id,
)
from src.gong_to_notion.notion_client import NotionClient

_ROOT = Path(__file__).resolve().parents[1]

HEADER_SENTINEL = "Account Name"


# ---------------------------------------------------------------------------
# Name normalization — matches xlsx Account Name to Notion Agency title
# ---------------------------------------------------------------------------

_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)
_WS_RE = re.compile(r"\s+")
_TRAILING_PAREN_RE = re.compile(r"\s*\(([^()]*)\)\s*$")


def normalize_name(raw: str) -> str:
    s = (raw or "").strip().lower()
    s = _PUNCT_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip()
    return s


def name_candidates(raw: str) -> set[str]:
    keys: set[str] = set()
    s = (raw or "").strip()
    if not s:
        return keys
    full = normalize_name(s)
    if full:
        keys.add(full)
    m = _TRAILING_PAREN_RE.search(s)
    if m:
        outer = _TRAILING_PAREN_RE.sub("", s).strip()
        inner = m.group(1).strip()
        if outer:
            k = normalize_name(outer)
            if k:
                keys.add(k)
        if inner:
            k = normalize_name(inner)
            if k:
                keys.add(k)
    return keys


# ---------------------------------------------------------------------------
# Column mapping config
# ---------------------------------------------------------------------------


# `xlsx_col` is the header text in the SFDC report. `notion_prop` is the
# Notion property name. `kind` controls payload shape and validation.
# Some xlsx columns appear with curly-quote variants — we normalize before
# header lookup, so the spec uses ASCII quotes.
@dataclass(frozen=True)
class ColMap:
    xlsx_col: str
    notion_prop: str
    kind: str  # title | text | select | multi_select | number | date


COLUMN_MAPPINGS: list[ColMap] = [
    ColMap("Account ID", "Account ID", "text"),
    ColMap("Account Stage", "Account Stage", "select"),
    ColMap("Market Segment", "CS Tier", "select"),
    ColMap("Customer ARR", "Customer ARR", "number"),
    ColMap("TAM", "TAM", "number"),
    ColMap("Contracted Fleet", "Contracted Fleet", "number"),
    ColMap("Product Summary", "Products", "multi_select"),
    ColMap("Customer Since", "Customer Since", "date"),
    ColMap(
        "Current CAD/AVL Contract Expiration",
        "Current CAD/AVL Contract Expiration",
        "date",
    ),
    ColMap("Total User Accounts", "User Accounts", "number"),
    ColMap("Renewal Date", "Renewal Date", "date"),
    ColMap("Churn Likelihood", "Churn Likelihood", "select"),
    ColMap("Renewal Risk Reason", "Renewal Risk Reason", "text"),
    ColMap("Who provides alerts for the agency?", "Who Provides Alerts", "select"),
    ColMap(
        "Passenger-Facing Tools Integrated",
        "Passenger Facing Tools",
        "multi_select",
    ),
    ColMap("Vehicle Health System Vendor", "Vehicle Health System", "select"),
    ColMap("Hardware Maintained by Swiftly", "Hardware Maintained by Swiftly", "select"),
    ColMap("Tablets", "Tablets", "select"),
    ColMap("APC Processor", "APC Processor", "select"),
    ColMap(
        "Partners integrated with Swiftly's APIs",
        "Partners Integrated with Swiftly API",
        "select",
    ),
    ColMap("Current CAD / AVL Vendor", "CAD/AVL System", "select"),
    ColMap("Current RTPI Provider", "Current RTPI Provider", "select"),
    ColMap("Scheduling Software", "Scheduling Software", "select"),
    ColMap("Organizational Type", "Organizational Type", "select"),
    ColMap("Website", "Website", "url"),
    ColMap("Number of Routes", "Routes", "number"),
    ColMap("Fixed Route Fleet Size", "Fleet Size", "number"),
]

# Non-ColMap column consumed by the parent-linking second pass.
PARENT_ID_COL = "Parent Account ID"


# ---------------------------------------------------------------------------
# Value transforms
# ---------------------------------------------------------------------------


_MARKET_SEGMENT_PREFIX_RE = re.compile(r"^\s*\d+\s*[\.\)\-]?\s*")


def transform_market_segment(raw: str) -> str:
    """SFDC values like '2.Enterprise' → 'Enterprise'; map 'Mid-Market' →
    'Mid-market' to match the existing Notion option."""
    s = _MARKET_SEGMENT_PREFIX_RE.sub("", raw or "").strip()
    # Normalize Mid-Market casing variants
    if s.lower() == "mid-market":
        return "Mid-market"
    return s


# Per-column value transform (raw cell str -> str). Default: identity.
COLUMN_TRANSFORMS: dict[str, callable] = {
    "Market Segment": transform_market_segment,
}


_SF_ID_SUFFIX_CHARS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ012345"


def sf_id_15_to_18(sf_id_15: str) -> str:
    """Standard Salesforce 15→18 char ID conversion (case-safe checksum)."""
    suffix = ""
    for chunk_start in (0, 5, 10):
        chunk = sf_id_15[chunk_start : chunk_start + 5]
        bits = sum(
            1 << i for i, ch in enumerate(chunk) if ch.isalpha() and ch.isupper()
        )
        suffix += _SF_ID_SUFFIX_CHARS[bits]
    return sf_id_15 + suffix


def sf_account_url(sf_id_15: str) -> str:
    return (
        "https://goswiftly.lightning.force.com/lightning/r/Account/"
        f"{sf_id_15_to_18(sf_id_15)}/view"
    )


def bucket_classification(stage: str) -> str | None:
    """Map raw SF Account Stage to the Notion `Classification` bucket."""
    if not stage:
        return None
    s = stage.strip().lower()
    if s == "won":
        return "Customer"
    if s in ("lost/nurture", "lost", "closed lost"):
        return "Closed Lost"
    return "Prospect"


# ---------------------------------------------------------------------------
# Date parsing (SF report uses M/D/YYYY)
# ---------------------------------------------------------------------------


def parse_date(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_number(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def split_multi(raw: str) -> list[str]:
    """Split a SF multi-value cell on commas or semicolons (SF mixes both),
    dropping empties and trailing punctuation."""
    if not raw:
        return []
    parts = re.split(r"[;,]", raw)
    return [p.strip().rstrip(".;,") for p in parts if p.strip().strip(",.;")]


# ---------------------------------------------------------------------------
# xlsx parsing
# ---------------------------------------------------------------------------


def _normalize_header_cell(c: Any) -> str:
    """Normalize curly quotes / extra whitespace so header lookup is forgiving."""
    s = "" if c is None else str(c)
    return (
        s.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .strip()
    )


def parse_accounts_xlsx_full(path: Path) -> tuple[list[dict[str, Any]], dict[str, int], list[str]]:
    """Return (rows, col_to_idx, owner_names_seen).

    Each row is a dict keyed by xlsx column name (only for columns we care
    about), with raw cell values (str/number/datetime). The xlsx leading
    title/filter rows are skipped via the 'Account Name' header sentinel.
    """
    import openpyxl  # type: ignore[import-not-found]

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx: int | None = None
    header: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [_normalize_header_cell(c) for c in row]
        if HEADER_SENTINEL in cells:
            header_row_idx = i
            header = cells
            break
    if header_row_idx is None:
        raise SystemExit(
            f"Could not find header row containing {HEADER_SENTINEL!r} in {path}"
        )

    # Index columns we recognize
    col_to_idx: dict[str, int] = {}
    expected = [
        HEADER_SENTINEL,
        "Account Owner",
        PARENT_ID_COL,
        *(m.xlsx_col for m in COLUMN_MAPPINGS),
    ]
    for col in expected:
        if col in header:
            col_to_idx[col] = header.index(col)
    missing = [c for c in expected if c not in col_to_idx]
    if missing:
        print(
            f"[xlsx] WARN: expected columns missing from export (skipped): "
            f"{', '.join(missing)}",
            file=sys.stderr,
        )

    # Account Stage is required for bucketing — fail loudly if missing.
    if "Account Stage" not in col_to_idx:
        raise SystemExit(
            "xlsx is missing the 'Account Stage' column. Re-export the SFDC "
            "Accounts Report with that column included."
        )
    if "Account ID" not in col_to_idx:
        raise SystemExit("xlsx is missing the 'Account ID' column.")

    rows: list[dict[str, Any]] = []
    owner_names: list[str] = []
    seen_names: set[str] = set()

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name_cell = row[col_to_idx[HEADER_SENTINEL]] if col_to_idx[HEADER_SENTINEL] < len(row) else None
        name = ("" if name_cell is None else str(name_cell)).strip()
        raw_id = row[col_to_idx["Account ID"]] if col_to_idx["Account ID"] < len(row) else None
        sf_id = normalize_sf_account_id(("" if raw_id is None else str(raw_id)).strip())
        if not name or not sf_id:
            continue
        if name in seen_names:
            print(
                f"[xlsx] WARN: duplicate Account Name {name!r} — keeping first",
                file=sys.stderr,
            )
            continue
        seen_names.add(name)

        record: dict[str, Any] = {"Account Name": name, "Account ID": sf_id}
        for col, idx in col_to_idx.items():
            if col in (HEADER_SENTINEL, "Account ID"):
                continue
            v = row[idx] if idx < len(row) else None
            if col == PARENT_ID_COL:
                v = normalize_sf_account_id(("" if v is None else str(v)).strip())
            record[col] = v

        if record.get("Account Owner"):
            owner_names.append(str(record["Account Owner"]).strip())
        rows.append(record)

    return rows, col_to_idx, owner_names


# ---------------------------------------------------------------------------
# Notion read helpers
# ---------------------------------------------------------------------------


def _read_title(prop: dict | None) -> str:
    if not prop:
        return ""
    parts = prop.get("title") or []
    return "".join((p.get("plain_text") or "") for p in parts).strip()


def _read_rich_text(prop: dict | None) -> str:
    if not prop:
        return ""
    parts = prop.get("rich_text") or []
    return "".join((p.get("plain_text") or "") for p in parts).strip()


def _read_select(prop: dict | None) -> str:
    if not prop:
        return ""
    sel = prop.get("select")
    return (sel or {}).get("name") or ""


def _read_multi_select(prop: dict | None) -> list[str]:
    if not prop:
        return []
    return [o.get("name") or "" for o in (prop.get("multi_select") or []) if o.get("name")]


def _read_url(prop: dict | None) -> str:
    if not prop:
        return ""
    return (prop.get("url") or "").strip()


def _read_relation_ids(prop: dict | None) -> list[str]:
    if not prop:
        return []
    return [r.get("id") or "" for r in (prop.get("relation") or []) if r.get("id")]


def _read_number(prop: dict | None) -> float | None:
    if not prop:
        return None
    return prop.get("number")


def _read_date(prop: dict | None) -> str:
    if not prop:
        return ""
    d = prop.get("date") or {}
    return d.get("start") or ""


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, list):
        return len(value) == 0
    return False


# ---------------------------------------------------------------------------
# Notion write payloads
# ---------------------------------------------------------------------------


def _title_payload(value: str) -> dict:
    return {"title": [{"type": "text", "text": {"content": value}}]}


def _rich_text_payload(value: str) -> dict:
    return {"rich_text": [{"type": "text", "text": {"content": value}}]}


def _select_payload(value: str) -> dict:
    return {"select": {"name": value}}


def _multi_select_payload(values: list[str]) -> dict:
    return {"multi_select": [{"name": v} for v in values]}


def _number_payload(value: float) -> dict:
    return {"number": value}


def _date_payload(iso: str) -> dict:
    return {"date": {"start": iso}}


def _url_payload(value: str) -> dict:
    return {"url": value}


def _relation_payload(page_ids: list[str]) -> dict:
    return {"relation": [{"id": pid} for pid in page_ids]}


# ---------------------------------------------------------------------------
# Schema lookup (option validation for select/multi_select)
# ---------------------------------------------------------------------------


@dataclass
class SchemaInfo:
    select_options: dict[str, set[str]] = field(default_factory=dict)
    multi_select_options: dict[str, set[str]] = field(default_factory=dict)
    prop_types: dict[str, str] = field(default_factory=dict)


def infer_schema_from_pages(pages: list[dict]) -> SchemaInfo:
    """Read every page's properties to learn the existing select/multi-select
    options. We avoid a separate schema endpoint by reading what's actually
    present on pages — sufficient for validation since we only validate
    against options Notion has seen before."""
    info = SchemaInfo()
    for p in pages:
        props = p.get("properties") or {}
        for name, prop in props.items():
            t = prop.get("type")
            if not t:
                continue
            info.prop_types.setdefault(name, t)
            if t == "select":
                opt = (prop.get("select") or {}).get("name")
                if opt:
                    info.select_options.setdefault(name, set()).add(opt)
            elif t == "multi_select":
                for o in prop.get("multi_select") or []:
                    if o.get("name"):
                        info.multi_select_options.setdefault(name, set()).add(o["name"])
    return info


# ---------------------------------------------------------------------------
# Per-row property build
# ---------------------------------------------------------------------------


@dataclass
class RowDelta:
    """Holds the desired property values for a single xlsx row, plus the
    select/multi-select values not yet seen as options in Notion (written
    anyway — Notion auto-creates options — but reported for tidy-up)."""
    name: str
    sf_id: str
    desired: dict[str, dict] = field(default_factory=dict)
    classification: str | None = None
    new_select_options: list[tuple[str, str]] = field(default_factory=list)  # (prop, value)
    new_multi_options: list[tuple[str, str]] = field(default_factory=list)


def _canon_option(known: set[str], value: str) -> str | None:
    """Return the existing option whose name matches case-insensitively."""
    lower = value.lower()
    for opt in known:
        if opt.lower() == lower:
            return opt
    return None


def build_row_delta(record: dict[str, Any], schema: SchemaInfo) -> RowDelta:
    name = str(record["Account Name"])
    sf_id = str(record["Account ID"])
    delta = RowDelta(name=name, sf_id=sf_id)

    for cm in COLUMN_MAPPINGS:
        if cm.xlsx_col not in record:
            continue
        raw = record.get(cm.xlsx_col)
        # Apply transform
        if cm.kind in ("select", "text") and isinstance(raw, str):
            tx = COLUMN_TRANSFORMS.get(cm.xlsx_col)
            if tx:
                raw = tx(raw)

        if cm.kind == "text":
            if raw is None:
                continue
            s = str(raw).strip()
            if not s:
                continue
            delta.desired[cm.notion_prop] = _rich_text_payload(s)

        elif cm.kind == "select":
            if raw is None:
                continue
            s = str(raw).strip()
            if not s or s.lower() == "no data":
                continue
            # Unknown options are written anyway (Notion auto-creates them);
            # case-insensitive hits reuse the existing option's casing so
            # 'INIT' doesn't create a near-dupe of 'Init'.
            known = schema.select_options.get(cm.notion_prop)
            if known is not None and s not in known:
                canon = _canon_option(known, s)
                if canon:
                    s = canon
                else:
                    delta.new_select_options.append((cm.notion_prop, s))
            delta.desired[cm.notion_prop] = _select_payload(s)

        elif cm.kind == "multi_select":
            if raw is None:
                continue
            parts = split_multi(str(raw))
            known = schema.multi_select_options.get(cm.notion_prop)
            kept: list[str] = []
            for p in parts:
                if p.lower() == "no data":
                    continue
                if known is not None and p not in known:
                    canon = _canon_option(known, p)
                    if canon:
                        p = canon
                    else:
                        delta.new_multi_options.append((cm.notion_prop, p))
                kept.append(p)
            if kept:
                delta.desired[cm.notion_prop] = _multi_select_payload(kept)

        elif cm.kind == "url":
            if raw is None:
                continue
            s = str(raw).strip()
            if not s:
                continue
            if not s.lower().startswith(("http://", "https://")):
                s = "https://" + s
            delta.desired[cm.notion_prop] = _url_payload(s)

        elif cm.kind == "number":
            n = parse_number(raw)
            if n is None:
                continue
            delta.desired[cm.notion_prop] = _number_payload(n)

        elif cm.kind == "date":
            iso = parse_date(raw)
            if not iso:
                continue
            delta.desired[cm.notion_prop] = _date_payload(iso)

    # Classification bucket from raw Account Stage
    stage_raw = record.get("Account Stage")
    if isinstance(stage_raw, str):
        bucket = bucket_classification(stage_raw)
        if bucket:
            delta.classification = bucket
            delta.desired["Classification"] = _select_payload(bucket)

    # Salesforce Link derived from the Account ID — no export column needed.
    delta.desired["Salesforce Link"] = _url_payload(sf_account_url(sf_id))

    return delta


# ---------------------------------------------------------------------------
# Diffing — blanks always fill; --update also overwrites differing values.
# xlsx blanks never reach `desired`, so nothing here can clear a Notion value.
# ---------------------------------------------------------------------------


def _url_norm(u: str) -> str:
    s = (u or "").strip().lower()
    for prefix in ("https://", "http://"):
        if s.startswith(prefix):
            s = s[len(prefix):]
    return s.rstrip("/")


def diff_properties(
    desired: dict[str, dict], page_props: dict, update_mode: bool
) -> tuple[dict[str, dict], list[tuple[str, str, str]], list[tuple[str, str, str]]]:
    """Return (updates_to_write, changes, conflicts).

    `changes` are (prop_name, xlsx_value, notion_value) overwrites included in
    `updates` because update_mode is on. `conflicts` are the same tuples when
    update_mode is off — differing values reported, never written.
    The page title is never diffed here; it's only used at create-time."""
    updates: dict[str, dict] = {}
    changes: list[tuple[str, str, str]] = []
    conflicts: list[tuple[str, str, str]] = []

    def resolve(prop_name: str, payload: dict, existing_blank: bool,
                differs: bool, new_repr: str, old_repr: str) -> None:
        if existing_blank:
            updates[prop_name] = payload
        elif differs:
            if update_mode:
                updates[prop_name] = payload
                changes.append((prop_name, new_repr, old_repr))
            else:
                conflicts.append((prop_name, new_repr, old_repr))

    for prop_name, payload in desired.items():
        existing = page_props.get(prop_name)

        if "rich_text" in payload:
            existing_val = _read_rich_text(existing)
            new_val = "".join(p["text"]["content"] for p in payload["rich_text"])
            resolve(prop_name, payload, _is_blank(existing_val),
                    existing_val != new_val, new_val, existing_val)
        elif "select" in payload:
            existing_val = _read_select(existing)
            new_val = (payload["select"] or {}).get("name") or ""
            resolve(prop_name, payload, _is_blank(existing_val),
                    existing_val != new_val, new_val, existing_val)
        elif "multi_select" in payload:
            existing_vals = _read_multi_select(existing)
            new_vals = [o["name"] for o in payload["multi_select"]]
            resolve(prop_name, payload, not existing_vals,
                    set(existing_vals) != set(new_vals),
                    ", ".join(new_vals), ", ".join(existing_vals))
        elif "number" in payload:
            existing_val = _read_number(existing)
            new_val = payload["number"]
            resolve(prop_name, payload, existing_val is None,
                    existing_val != new_val, str(new_val), str(existing_val))
        elif "date" in payload:
            existing_val = _read_date(existing)
            new_val = (payload["date"] or {}).get("start") or ""
            resolve(prop_name, payload, _is_blank(existing_val),
                    existing_val != new_val, new_val, existing_val)
        elif "url" in payload:
            existing_val = _read_url(existing)
            new_val = payload["url"]
            # Scheme/trailing-slash differences are noise, not changes.
            resolve(prop_name, payload, _is_blank(existing_val),
                    _url_norm(existing_val) != _url_norm(new_val),
                    new_val, existing_val)

    return updates, changes, conflicts


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    load_dotenv(_ROOT / ".env")
    token = os.getenv("NOTION_TOKEN")
    if not token:
        raise SystemExit("NOTION_TOKEN must be set in .env")

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="Path to the SFDC Accounts Report xlsx.")
    ap.add_argument(
        "--apply",
        action="store_true",
        help="Actually write to Notion. Default is dry-run.",
    )
    ap.add_argument(
        "--no-create",
        action="store_true",
        help="When --apply is set, only update existing Agencies; do not create new pages.",
    )
    ap.add_argument(
        "--update",
        action="store_true",
        help="Overwrite Notion values that differ from the xlsx (SF is the "
        "source of truth). Default only fills blanks and reports conflicts.",
    )
    ap.add_argument(
        "--create-stages",
        help="Comma-separated Account Stages; only unmatched rows in these "
        "stages get new pages (e.g. \"Won,Engaged,Early Stage\"). "
        "Default: all stages.",
    )
    args = ap.parse_args()

    create_stages: set[str] | None = None
    if args.create_stages:
        create_stages = {
            s.strip().lower() for s in args.create_stages.split(",") if s.strip()
        }

    if not args.xlsx.exists():
        raise SystemExit(f"File not found: {args.xlsx}")

    dry_run = not args.apply

    print(f"[xlsx] reading {args.xlsx} ...", file=sys.stderr)
    xlsx_rows, col_index, owner_names = parse_accounts_xlsx_full(args.xlsx)
    present_cols = sorted(col_index.keys())
    print(
        f"  {len(xlsx_rows)} rows; columns: {', '.join(present_cols)}",
        file=sys.stderr,
    )

    with NotionClient(token) as notion:
        print("[notion] loading Agencies ...", file=sys.stderr)
        agencies = notion.query_data_source(AGENCIES_DATA_SOURCE_ID)
        print(f"  {len(agencies)} agency pages", file=sys.stderr)

        schema = infer_schema_from_pages(agencies)

        # Index Notion pages by every candidate key derivable from Name and Full name,
        # plus by normalized SF Account ID (preferred when present — avoids name
        # collisions when SF has multiple accounts that share a display name).
        key_to_pages: dict[str, set[str]] = {}
        sf_id_to_page: dict[str, str] = {}
        # page_info[pid] = (display_name, props_dict)
        page_info: dict[str, tuple[str, dict]] = {}
        for a in agencies:
            pid = a.get("id")
            if not pid:
                continue
            props = a.get("properties") or {}
            name = _read_title(props.get("Name"))
            full = _read_rich_text(props.get("Full name"))
            page_info[pid] = (name, props)
            for k in name_candidates(name) | name_candidates(full):
                key_to_pages.setdefault(k, set()).add(pid)
            existing_sf = normalize_sf_account_id(_read_rich_text(props.get("Account ID")))
            if existing_sf:
                sf_id_to_page[existing_sf] = pid

        # Stats
        update_writes: list[tuple[str, str, dict]] = []  # (page_id, name, props_payload)
        already_set_count = 0
        prop_fill_counts: Counter = Counter()
        prop_change_counts: Counter = Counter()
        all_changes: list[tuple[str, str, str, str]] = []  # (name, prop, xlsx_val, notion_val)
        all_conflicts: list[tuple[str, str, str, str]] = []  # (name, prop, xlsx_val, notion_val)
        title_mismatches: list[tuple[str, str]] = []  # (notion_title, sf_name)
        ambiguous: list[tuple[str, list[str]]] = []
        unmatched_xlsx: list[dict[str, Any]] = []
        matched_pages: set[str] = set()
        record_page: dict[str, str] = {}  # xlsx Account ID → matched/created page id
        new_select_total: Counter = Counter()
        new_multi_total: Counter = Counter()
        bucket_counts: Counter = Counter()
        stage_counts: Counter = Counter()

        for record in xlsx_rows:
            stage_counts[str(record.get("Account Stage") or "").strip()] += 1
            delta = build_row_delta(record, schema)
            for p, v in delta.new_select_options:
                new_select_total[(p, v)] += 1
            for p, v in delta.new_multi_options:
                new_multi_total[(p, v)] += 1
            if delta.classification:
                bucket_counts[delta.classification] += 1

            # Prefer SF Account ID match — avoids false hits when SF has
            # multiple accounts sharing a display name.
            pid: str | None = sf_id_to_page.get(record["Account ID"])
            if pid is None:
                candidates = name_candidates(record["Account Name"])
                hits: set[str] = set()
                for k in candidates:
                    hits |= key_to_pages.get(k, set())
                # Drop any name hits that already have a *different* SF Account ID —
                # those are clearly different accounts that happen to share a name.
                hits = {
                    h for h in hits
                    if not (
                        normalize_sf_account_id(
                            _read_rich_text(page_info[h][1].get("Account ID"))
                        )
                        and normalize_sf_account_id(
                            _read_rich_text(page_info[h][1].get("Account ID"))
                        ) != record["Account ID"]
                    )
                }
                if not hits:
                    unmatched_xlsx.append(record)
                    continue
                if len(hits) > 1:
                    ambiguous.append(
                        (record["Account Name"], sorted(page_info[h][0] for h in hits))
                    )
                    continue
                (pid,) = hits
            matched_pages.add(pid)
            record_page[record["Account ID"]] = pid
            # Name-matched pages get their Account ID filled this run; index
            # them now so the parent-linking pass can resolve them.
            sf_id_to_page.setdefault(record["Account ID"], pid)
            notion_name, notion_props = page_info[pid]
            if notion_name.strip() != record["Account Name"].strip():
                title_mismatches.append((notion_name, record["Account Name"]))
            updates, changes, conflicts = diff_properties(
                delta.desired, notion_props, args.update
            )
            changed_props = {prop for prop, _, _ in changes}
            for prop, xv, nv in changes:
                all_changes.append((notion_name, prop, xv, nv))
                prop_change_counts[prop] += 1
            for prop, xv, nv in conflicts:
                all_conflicts.append((notion_name, prop, xv, nv))
            if updates:
                update_writes.append((pid, notion_name, updates))
                for prop in updates.keys():
                    if prop not in changed_props:
                        prop_fill_counts[prop] += 1
            else:
                already_set_count += 1

        unmatched_agencies = [
            (info[0], _read_rich_text(info[1].get("Account ID")))
            for pid, info in page_info.items()
            if pid not in matched_pages
        ]

        # Apply updates
        if not dry_run:
            for pid, notion_name, payload in update_writes:
                notion.update_page(pid, payload)
                print(f"  updated {notion_name!r}: {sorted(payload.keys())}", file=sys.stderr)

        # Apply creates
        creates_done: list[str] = []
        creates_planned: list[tuple[str, str]] = []  # (name, stage)
        creates_skipped_by_stage: Counter = Counter()
        would_create_ids: set[str] = set()  # xlsx Account IDs of planned creates (dry-run)
        if not args.no_create:
            for record in unmatched_xlsx:
                stage = str(record.get("Account Stage") or "").strip()
                if create_stages is not None and stage.lower() not in create_stages:
                    creates_skipped_by_stage[stage or "(blank)"] += 1
                    continue
                if dry_run:
                    would_create_ids.add(record["Account ID"])
                delta = build_row_delta(record, schema)
                props_payload: dict[str, dict] = {
                    "Name": _title_payload(record["Account Name"]),
                    **delta.desired,
                }
                creates_planned.append((record["Account Name"], stage or "(blank)"))
                if not dry_run:
                    created = notion.create_page(AGENCIES_DATA_SOURCE_ID, props_payload)
                    if created.get("id"):
                        record_page[record["Account ID"]] = created["id"]
                        sf_id_to_page[record["Account ID"]] = created["id"]
                    creates_done.append(record["Account Name"])
                    print(f"  created {record['Account Name']!r}", file=sys.stderr)

        # Second pass: Parent Agency links. Runs after creates so parents
        # created in this run resolve. In dry-run, planned creates count as
        # resolvable targets/children even though no page exists yet.
        parent_links_planned = 0
        parent_changes: list[tuple[str, str, str]] = []  # (child, old parent, new parent)
        parent_conflicts: list[tuple[str, str, str]] = []
        parent_unresolved: list[tuple[str, str]] = []  # (child name, parent sf id)
        for record in xlsx_rows:
            parent_sf = record.get(PARENT_ID_COL) or ""
            if not parent_sf or parent_sf == record["Account ID"]:
                continue
            child_name = record["Account Name"]
            parent_pid = sf_id_to_page.get(parent_sf)
            parent_is_pending = dry_run and parent_sf in would_create_ids
            if not parent_pid and not parent_is_pending:
                parent_unresolved.append((child_name, parent_sf))
                continue
            child_pid = record_page.get(record["Account ID"])
            child_is_pending = dry_run and record["Account ID"] in would_create_ids
            if not child_pid and not child_is_pending:
                continue  # child row neither matched nor being created
            # Freshly created children (and dry-run pending ones) have no
            # existing relation; matched pages might.
            existing_parents: list[str] = []
            if child_pid and child_pid in page_info:
                existing_parents = _read_relation_ids(
                    page_info[child_pid][1].get("Parent Agency")
                )
            if not existing_parents:
                parent_links_planned += 1
                if not dry_run and child_pid and parent_pid:
                    notion.update_page(
                        child_pid, {"Parent Agency": _relation_payload([parent_pid])}
                    )
            elif parent_pid and existing_parents != [parent_pid]:
                old_names = ", ".join(
                    page_info.get(p, ("<unknown>",))[0] for p in existing_parents
                )
                new_name = page_info.get(parent_pid, ("<new page>",))[0]
                if args.update:
                    parent_changes.append((child_name, old_names, new_name))
                    if not dry_run and child_pid:
                        notion.update_page(
                            child_pid,
                            {"Parent Agency": _relation_payload([parent_pid])},
                        )
                else:
                    parent_conflicts.append((child_name, old_names, new_name))

    # -------- Report --------
    print("")
    print("=" * 78)
    print(
        f"{'DRY-RUN — ' if dry_run else ''}Agency upsert from xlsx"
        f"{' (--update: overwrites enabled)' if args.update else ' (blanks-only)'}"
    )
    print("=" * 78)
    print(f"  xlsx rows:           {len(xlsx_rows)}")
    print(f"  matched (existing):  {len(matched_pages)}")
    print(f"    fully up-to-date:  {already_set_count}")
    print(
        f"    {'would update' if dry_run else 'updated'}:      "
        f"{len(update_writes)}"
    )
    if args.update:
        print(f"    value overwrites:  {len(all_changes)}  (across those pages)")
    print(f"  unmatched xlsx:      {len(unmatched_xlsx)}  (no Agency with that name)")
    if args.no_create:
        print(f"    {'would create' if dry_run else 'created'}:      0  (--no-create)")
    else:
        print(
            f"    {'would create' if dry_run else 'created'}:      "
            f"{len(creates_planned) if dry_run else len(creates_done)}"
        )
        if creates_skipped_by_stage:
            skipped_n = sum(creates_skipped_by_stage.values())
            print(f"    skipped by stage:  {skipped_n}  (--create-stages filter)")
    print(f"  ambiguous matches:   {len(ambiguous)}")
    if not args.update:
        print(f"  conflicts:           {len(all_conflicts)}  (existing value differs — NOT overwritten)")
    print(
        f"  parent links:        {parent_links_planned} "
        f"{'planned' if dry_run else 'written'}, "
        f"{len(parent_changes)} changed, {len(parent_conflicts)} conflicts, "
        f"{len(parent_unresolved)} unresolved"
    )
    print(f"  title mismatches:    {len(title_mismatches)}  (reported only, never written)")
    print(
        f"  unmatched Notion:    {len(unmatched_agencies)}  "
        f"(Agencies not present in xlsx)"
    )

    print("\nAccount Stage distribution (xlsx):")
    for stage, n in stage_counts.most_common():
        print(f"  {n:5}  {stage!r}")

    print("\nClassification bucket counts (derived):")
    for b, n in bucket_counts.most_common():
        print(f"  {n:5}  {b}")

    if prop_fill_counts:
        print(f"\nBlank-property fills on existing pages ({'planned' if dry_run else 'applied'}):")
        for prop, n in prop_fill_counts.most_common():
            print(f"  {n:5}  {prop}")

    if prop_change_counts:
        print(f"\nValue overwrites by property ({'planned' if dry_run else 'applied'} — --update):")
        for prop, n in prop_change_counts.most_common():
            print(f"  {n:5}  {prop}")

    if all_changes:
        # Classification downgrades off Customer are the highest-stakes
        # overwrites — always show every one.
        downgrades = [
            c for c in all_changes if c[1] == "Classification" and c[3] == "Customer"
        ]
        if downgrades:
            print(f"\n!! Customer downgrades ({len(downgrades)} — verify these are real):")
            for name, prop, xv, nv in downgrades:
                print(f"  - {name!r}  Classification: {nv!r} → {xv!r}")

        print(f"\nValue overwrites ({'planned' if dry_run else 'applied'}; notion → xlsx):")
        for name, prop, xv, nv in all_changes[:80]:
            print(f"  - {name!r}  {prop}: {nv!r} → {xv!r}")
        if len(all_changes) > 80:
            print(f"  ... and {len(all_changes) - 80} more")

        diff_path = Path("agency_upsert_changes.tsv")
        with diff_path.open("w") as f:
            f.write("agency\tproperty\tnotion_value\txlsx_value\n")
            for name, prop, xv, nv in all_changes:
                f.write(f"{name}\t{prop}\t{nv}\t{xv}\n")
        print(f"\nFull overwrite list ({len(all_changes)} rows): {diff_path.resolve()}")

    if new_select_total:
        print("\nNew select options ({} — review colors/dupes in Notion):".format(
            "will be created" if dry_run else "created"
        ))
        for (prop, val), n in new_select_total.most_common():
            print(f"  {n:5}  {prop} = {val!r}")

    if new_multi_total:
        print("\nNew multi-select options ({}):".format(
            "will be created" if dry_run else "created"
        ))
        for (prop, val), n in new_multi_total.most_common():
            print(f"  {n:5}  {prop} contains {val!r}")

    if all_conflicts:
        print("\nConflicts (existing Notion value differs from xlsx — NOT overwritten; use --update):")
        for name, prop, xv, nv in all_conflicts[:50]:
            print(f"  - {name!r}  {prop}: xlsx={xv!r}  notion={nv!r}")
        if len(all_conflicts) > 50:
            print(f"  ... and {len(all_conflicts) - 50} more")

    if title_mismatches:
        print("\nTitle mismatches (never written — rename by hand if warranted):")
        for notion_title, sf_name in title_mismatches:
            print(f"  - notion={notion_title!r}  sf={sf_name!r}")

    if creates_planned and dry_run:
        print("\nWould-create pages by Account Stage:")
        by_stage: dict[str, list[str]] = {}
        for name, stage in creates_planned:
            by_stage.setdefault(stage, []).append(name)
        for stage in sorted(by_stage, key=lambda s: -len(by_stage[s])):
            names = sorted(by_stage[stage])
            print(f"  {stage} ({len(names)}):")
            for n in names:
                print(f"    - {n}")

    if creates_skipped_by_stage:
        print("\nUnmatched rows skipped by --create-stages:")
        for stage, n in creates_skipped_by_stage.most_common():
            print(f"  {n:5}  {stage}")

    if parent_changes:
        print("\nParent Agency overwrites (--update; old → new):")
        for child, old, new in parent_changes:
            print(f"  - {child!r}: {old!r} → {new!r}")

    if parent_conflicts:
        print("\nParent Agency conflicts (existing differs — NOT overwritten; use --update):")
        for child, old, new in parent_conflicts:
            print(f"  - {child!r}: notion={old!r}  xlsx={new!r}")

    if parent_unresolved:
        print(f"\nParent Account IDs with no matching Agency ({len(parent_unresolved)}):")
        for child, parent_sf in parent_unresolved[:30]:
            print(f"  - {child!r} → parent SF {parent_sf}")
        if len(parent_unresolved) > 30:
            print(f"  ... and {len(parent_unresolved) - 30} more")

    if ambiguous:
        print("\nAmbiguous matches (xlsx name matches multiple Agencies — resolve by hand):")
        for xlsx_name, notion_names in ambiguous:
            print(f"  - {xlsx_name!r} ↔ {notion_names}")

    if owner_names:
        owner_counts = Counter(owner_names)
        print(f"\nDistinct Account Owners in xlsx ({len(owner_counts)} unique — populate Account Manager by hand):")
        for o, n in owner_counts.most_common():
            print(f"  {n:5}  {o}")

    if unmatched_agencies:
        print(
            f"\nNotion Agencies not in xlsx ({len(unmatched_agencies)} — left untouched, "
            "expected because xlsx is filtered to recent activity):"
        )
        for name, current in sorted(unmatched_agencies)[:20]:
            suffix = f"  [Account ID: {current}]" if current else ""
            print(f"  - {name}{suffix}")
        if len(unmatched_agencies) > 20:
            print(f"  ... and {len(unmatched_agencies) - 20} more")

    if dry_run:
        print("\n[DRY-RUN] No writes performed. Re-run with --apply to write.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
